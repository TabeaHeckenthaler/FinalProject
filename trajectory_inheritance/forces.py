from Directories import excel_sheet_directory
from openpyxl import load_workbook
from os import path
import numpy as np
import csv
import pandas as pd
from scipy import stats
from Setup.Maze import Maze
from PhysicsEngine.drawables import Arrow
from scipy.signal import find_peaks


def get_sheet():
    workbook = load_workbook(filename=excel_sheet_directory + path.sep + "Testable.xlsx")
    sheet = workbook.active
    return sheet


sheet = get_sheet()
DISPLAY_CONSTANT = 0.2


class Forces:
    def __init__(self, humans, x):
        from trajectory_inheritance.humans import participant_number
        self.excel_index = humans.excel_index
        if self.get_force_filename() is not None:
            self.date = self.get_date()
            self.size = humans.size
            self.directory = self.force_directory()
            self.occupied = humans.occupied
            self.filename = self.get_force_filename()
            self.abs_values = self.forces_loading(humans.frames, x.fps)
            self.angles = self.get_angles(humans, x)
            self.angles_load = self.angles - x.angle[:, np.newaxis]
            self.force_meters = Maze(x).force_attachment_positions_in_trajectory(x, reference_frame='load')

    @staticmethod
    def get_angles(humans, x):
        """
        :param humans: object of the class Humans
        :param x: object of the class trajectory_inheritance.trajectory
        :return: angles of the forces in world coordinates
        """
        from trajectory_inheritance.humans import angle_shift
        angle_shift = np.array([angle_shift[x.size][i]
                                for i in range(len(angle_shift[x.size].keys()))])[np.newaxis, :]
        return humans.angles + x.angle[:, np.newaxis] + angle_shift

    def get_date(self):
        day = sheet.cell(row=self.excel_index, column=2).value
        daytime = sheet.cell(row=self.excel_index, column=6).value
        return day.replace(hour=daytime.hour, minute=daytime.minute)

    def force_directory(self):
        day_string = str(self.date.year) + '-' + str(self.date.month).zfill(2) + '-' + str(self.date.day).zfill(2)
        return ('{0}{1}phys-guru-cs{2}ants{3}Tabea{4}Human Experiments{5}Raw Data and Videos{7}'
                + day_string + '{8}Force Measurements{9}' + self.size).format(path.sep, path.sep, path.sep, path.sep,
                                                                              path.sep, path.sep, path.sep, path.sep,
                                                                              path.sep, path.sep)

    def synchronization_offset(self, fps: int):
        """
        :param fps: frames per second
        If there is no force meter measurement return None.
        :return: frame of turning on force meter relative to start of the raw movie
        """
        if sheet.cell(row=self.excel_index, column=16).value == '/':
            return None

        if sheet.cell(row=self.excel_index, column=16).value is None:
            raise Exception('Fill in the Force synchronization time in line ' + str(self.excel_index))

        [minute, second] = [int(number) for number in
                            sheet.cell(row=self.excel_index, column=16).value.strip()[:-3].split(':')]
        frame_force_meter = (second + minute * 60) * fps

        """ if the frame of synchronization is BEFORE the start of the movie which was tracked """
        if sheet.cell(row=self.excel_index, column=16).value[0] == '-':
            frame_force_meter = - frame_force_meter

        """ time of tracking relative to start of the raw movie """
        raw_string = sheet.cell(row=self.excel_index, column=8).value
        if ', ' in raw_string:
            frame_tracking = int(raw_string.split(', ')[0])
        else:
            frame_tracking = int(raw_string.split('\n')[0])
        return frame_tracking - frame_force_meter

    def get_force_filename(self):
        txt_name = sheet.cell(row=self.excel_index, column=19).value
        if txt_name.endswith('.txt') or txt_name.endswith('.TXT'):
            return txt_name
        elif txt_name == '/':
            return None
        else:
            raise ValueError('You still have to add the name of the force file in line ' + str(self.excel_index))

    def forces_loading(self, frames, fps):
        from trajectory_inheritance.humans import participant_number
        # read force meter file
        with open(self.force_directory() + path.sep + self.filename, 'r') as f:
            reader = csv.reader(f, delimiter='\t')
            text_file_content = [line for line in reader]

        def convert_to_frames(fps, times):
            def correct_times():
                # for the large human SPT Maze, we have an additional two digits and an space in our txt file.
                # We have to get rid of this.
                for i, time in enumerate(times):
                    times[i] = [times[i][0].split(' ')[-1]]
                return times

            times = correct_times()
            seconds = [int(time[0].split(':')[0]) * 3600 + int(time[0].split(':')[1]) * 60 + int(time[0].split(':')[2])
                       for time
                       in times]
            seconds = [sec - seconds[0] for sec in seconds]
            frames = []
            for second in range(seconds[-1]):
                measurements_per_second = len([sec for sec in seconds if sec == second])
                for ii in range(measurements_per_second):
                    frames.append(second * fps + int(ii * fps / measurements_per_second))
            return frames

        sampled_frames = convert_to_frames(fps, text_file_content[1::2][1:-1])

        forces_txt = [[float(fu) for fu in fo[0].split(' ') if len(fu) > 1] for fo in text_file_content[0::2][:-1]]

        # all unoccupied force meters should have zero force
        empty_indices = [i for i in range(participant_number[self.size]) if i not in self.occupied]
        for empty_index in empty_indices:
            for j in range(len(forces_txt)):
                forces_txt[j][empty_index] = 0

        # every frame of the movie gets a force for every force meter
        forces_all_frames = []
        for frames_index in range(len(sampled_frames) - 1):
            for ii in range(sampled_frames[frames_index], sampled_frames[frames_index + 1]):
                forces_all_frames.append(forces_txt[frames_index])

        # find the offset of the first frame of the movie to the start of the force meter measurement
        synch_offset = self.synchronization_offset(fps)

        # write the force into the self.frame[:].forces variable
        if len(forces_all_frames) < len(frames) + synch_offset:
            if sheet.cell(row=self.excel_index, column=18).value == None:
                print()
            if 'battery' in sheet.cell(row=self.excel_index, column=18).value:
                print('Battery empty')
                empty = [0.0 for _ in range(len(forces_all_frames[0]))]
                missing_frames = range(-len(forces_all_frames) + (len(frames) + synch_offset + 10))
                [forces_all_frames.append(empty) for _ in missing_frames]

        abs_values = []
        for i, force_index in enumerate(range(synch_offset, len(frames) + synch_offset)):
            abs_values.append(forces_all_frames[force_index])

        abs_values = np.array(abs_values)

        # in some cases, forces are suddenly negative
        for i in range(abs_values.shape[1]):
            if len(np.where(abs_values[:, i] < 0)[0]) / abs_values.shape[0] > 0.6:
                abs_values[:, i] = -abs_values[:, i]

        abs_values = self.remove_force_outliers(np.array(abs_values))
        abs_values = abs_values - self.plateaus(abs_values)
        return abs_values

    @staticmethod
    def remove_force_outliers(array):
        def remove_force_outliers_single_forcemeter(single):
            # only one measurement
            df_original = pd.DataFrame(single)

            outlier_index = np.where((np.abs(stats.zscore(df_original, axis=0)) < 5) == False)[0]
            df_original.values[outlier_index] = 0 # TODO: this should be NaN, .. I think
            df_no_outliers = df_original.interpolate()
            return df_no_outliers
        return np.squeeze(np.apply_along_axis(remove_force_outliers_single_forcemeter, 0, array))

    @staticmethod
    def plateaus(arrays):

        def plateau(array):
            plateaus = find_peaks(array, plateau_size=20)[0]
            if len(plateaus) == 0:
                return array.min()
            if len(np.where(array - array[plateaus].mean() < 0)[0]) / len(array) > 0.4:
                return np.nanmin(array)
            return array[plateaus].mean()

        return [plateau(arrays[:, i]) for i in range(arrays.shape[1])]

    def draw(self, display, x):
        force_attachments = display.my_maze.force_attachment_positions()
        for name in x.participants.occupied:
            self.arrow(display.i, force_attachments[name], name).draw(display)

    def torque(self, part):
        return np.cross(self.force_vector(part, reference_frame='load'), self.force_meters[:, part])

    def arrow(self, i, force_meter_coor, name) -> Arrow:
        """
        :param i: index of the participant
        :param force_meter_coor: where is the force_meter located in world coordinates
        :return: start, end and string for the display of the force as a triplet
        """
        start = force_meter_coor
        end = force_meter_coor + self.abs_values[i, name] * DISPLAY_CONSTANT * \
              np.array([np.cos(self.angles[i, name]), np.sin(self.angles[i, name])])
        return Arrow(np.array(start), np.array(end), str(name))

    def force_vector(self, name: int, reference_frame='maze') -> np.ndarray:
        """
        :param name: index of the participant
        :param reference_frame: 'maze' or 'load', dependent on desired reference frame
        :return: len(x.frames)x2 numpy.array with x and y components of the force vectors
        """
        if reference_frame == 'maze':
            a = self.angles[:, name]
        elif reference_frame == 'load':
            a = self.angles_load[:, name]
        else:
            raise ValueError('What frame of reference?')

        return np.transpose(np.array([np.cos(a), np.sin(a)]) * self.abs_values[:, name])

    # def debugger(human, forces_all_frames, x):
    #     if np.isnan(np.sum([human.frames[i].forces[1] for i in range(0, len(human.frames))])):
    #         with open(excel_sheet_directory + path.sep + 'mistakes.txt', 'a') as f:
    #             f.write('\n' + x.filename + '\n')
    #             f.write('original movie: ' + sheet.cell(row=human.excel_index, column=1).value + '\n')
    #             f.write('force file: ' + force_filename(human) + '\n')
    #             f.write(
    #                 'configuration time: ' + sheet.cell(row=human.excel_index, column=16).value + '\n')
    #             f.write('length of force measurement: ' +
    #                     str(int(np.floor(len(forces_all_frames) / x.fps / 60))).zfill(2) + ':' +
    #                     str(int(np.floor(len(forces_all_frames) / x.fps % 60))).zfill(2) + '\n')
    #             f.write('missing frames: ' +
    #                     str(len([i for i in range(len(human.frames))
    #                              if np.isnan(human.frames[i].forces[0])])) + '\n')
